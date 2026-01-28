import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

file_path="C:\\Users\\admin\\Downloads\\download.xlsx"
fruit = "Apple";
colName = "price"
new_value = "700";


def updateExcelData(file_path, fruit, colName, new_value):
    book = openpyxl.load_workbook(file_path);
    sheet = book.active;
    dict = {}
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            dict["column"] = i;

    for i in range(1, sheet.max_row + 1):  # Max col +1
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == fruit:
                print("before update: " + str(sheet.cell(row=i, column=j).value));
                dict["row"] = i
                break
    sheet.cell(row=dict["row"], column=dict["column"]).value = new_value;
    book.save(file_path)


driver = webdriver.Chrome()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html");
driver.implicitly_wait(5);
driver.find_element(By.ID, "downloadButton").click();
updateExcelData(file_path, fruit, colName, new_value);
#Download button
# input[type=file]
file_input=driver.find_element(By.CSS_SELECTOR, "input[type='file']");
file_input.send_keys(file_path);
wait = WebDriverWait(driver, 10);
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)");
wait.until(expected_conditions.visibility_of_element_located(toast_locator));
print(driver.find_element(*toast_locator).text)

column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
price=driver.find_element(By.XPATH, "//div[text()='"+fruit+"']/parent::div/parent::div/div[@id='cell-"+str(column)+"-undefined']")
print(price.text);
assert price.text == new_value;


