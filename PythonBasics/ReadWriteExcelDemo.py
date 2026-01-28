#use openpyxl

import openpyxl

book = openpyxl.load_workbook("C:\\Users\\admin\\PycharmProjects\\PythonProject\\Selenium_Python_TestData.xlsx");
sheet = book.active;
Dictionary = {}

#Read excel file
cell = sheet.cell(row=2, column=2);
print(cell.value)

#write
sheet.cell(row=17, column=2).value = "16"
sheet.cell(row=17, column=3).value = "Johanna"
print(" Print using cell :" + sheet["B17"].value);
print(" Print using cell :" + sheet["C17"].value);

for i in range(2,sheet.max_row+1):#Max col +1
    for j in range(2, sheet.max_column+1):
        if sheet.cell(row=i, column=3).value =="James":
            print(sheet.cell(row=i, column=j).value);
            #dictionary["name"] = 'shetty'
            Dictionary[sheet.cell(row=1,column=j).value]= sheet.cell(row=i, column=j).value
print(Dictionary);

#Store data in dictionary